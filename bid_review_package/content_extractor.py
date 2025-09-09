import requests
import os
from urllib.parse import urlparse
from PyPDF2 import PdfReader
from docx import Document
import pytesseract
from PIL import Image
import io
import openpyxl  # 新增：用于Excel
import json      # 新增：用于JSON
import xml.etree.ElementTree as ET  # 新增：用于XML
from sqlalchemy import create_engine, text  # 新增：用于数据库
import pymongo  # 新增：用于MongoDB

def download_file(url):
    response = requests.get(url)
    response.raise_for_status()  # Raise an error for bad responses
    return response.content

def extract_from_pdf(content):
    pdf_file = io.BytesIO(content)
    reader = PdfReader(pdf_file)
    text = ""
    for page in reader.pages:
        text += page.extract_text() + "\n"
    return text.strip()

def extract_from_word(content, ext=".docx"):
    doc_file = io.BytesIO(content)
    doc = Document(doc_file)
    text = ""
    for para in doc.paragraphs:
        text += para.text + "\n"
    return text.strip()

def extract_from_image(content):
    img_file = io.BytesIO(content)
    img = Image.open(img_file)
    text = pytesseract.image_to_string(img)
    return text.strip()

# 新增：提取Excel内容
def extract_from_excel(content):
    excel_file = io.BytesIO(content)
    workbook = openpyxl.load_workbook(excel_file)
    text = ""
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    text += str(cell) + " "
            text += "\n"
    return text.strip()

# 新增：提取JSON内容
def extract_from_json(content):
    json_str = content.decode('utf-8')
    data = json.loads(json_str)
    return json.dumps(data, indent=4, ensure_ascii=False)

# 新增：提取XML内容
def extract_from_xml(content):
    xml_str = content.decode('utf-8')
    root = ET.fromstring(xml_str)
    text = ""
    for elem in root.iter():
        if elem.text and elem.text.strip():
            text += elem.text.strip() + "\n"
    return text.strip()

# 新增：提取数据库内容
def extract_from_database(url):
    try:
        if 'mongodb' in url:
            # MongoDB
            client = pymongo.MongoClient(url)
            db = client.get_default_database()
            text = ""
            for collection_name in db.list_collection_names():
                collection = db[collection_name]
                for doc in collection.find():
                    text += json.dumps(doc, indent=4, ensure_ascii=False) + "\n"
            return text.strip()
        else:
            # SQL数据库 (MySQL, PostgreSQL等)
            engine = create_engine(url)
            text = ""
            with engine.connect() as conn:
                if 'sqlite' in url:
                    result = conn.execute(text("SELECT name FROM sqlite_master WHERE type='table'"))
                else:
                    result = conn.execute(text("SHOW TABLES"))  # 假设MySQL/PostgreSQL
                tables = [row[0] for row in result]
                for table in tables:
                    result = conn.execute(text(f"SELECT * FROM {table} LIMIT 10"))  # 限制行数
                    columns = result.keys()
                    text += f"Table: {table}\n"
                    for row in result:
                        text += str(dict(zip(columns, row))) + "\n"
                    text += "\n"
                return text.strip()
    except Exception as e:
        return f"Error extracting from database: {str(e)}"

def read_local_file(path: str) -> bytes:
    """读取本地文件内容"""
    with open(path, "rb") as f:
        return f.read()

def extract_content(source: str):
    try:
        # 判断是数据库连接字符串
        low = source.lower()
        if any(proto in low for proto in (
            "mysql://", "mysql+pymysql://", "postgresql://", "postgresql+psycopg2://",
            "sqlite://", "sqlite+pysqlite://", "mongodb://", "mongodb+srv://"
        )):
            return extract_from_database(source)

        # 判断是 URL 还是本地文件
        if source.startswith("http://") or source.startswith("https://"):
            content, content_type = download_file(source)
            parsed = urlparse(source)
            ext = os.path.splitext(parsed.path)[1].lower()
        else:
            content = read_local_file(source)
            ext = os.path.splitext(source)[1].lower()
            content_type = None

        # JSON 特殊判断（本地 .json 或 API 返回）
        if ext == ".json" or "api.github.com" in low:
            return extract_from_json(content)

        if ext == ".pdf":
            return extract_from_pdf(content)
        elif ext in (".docx", ".doc"):
            return extract_from_word(content, ext)
        elif ext in (".png", ".jpg", ".jpeg", ".bmp", ".tiff"):
            return extract_from_image(content)
        elif ext in (".xlsx", ".xls"):
            return extract_from_excel(content)
        elif ext == ".xml":
            return extract_from_xml(content)
        elif ext == ".txt":
            return content.decode("utf-8", errors="ignore")
        else:
            return (f"[WARN] 未识别或不支持的文件类型。\n"
                    "支持：PDF、DOCX、PNG/JPG/BMP/TIFF、XLSX/XLS、JSON、XML、TXT、"
                    "以及数据库URL（MySQL/PostgreSQL/SQLite/MongoDB）。\n"
                    f"扩展名: {ext or '无'} Content-Type: {content_type or '未知'}")
    except Exception as e:
        return f"[ERROR] 提取失败: {e}"

if __name__ == "__main__":
    url = input("Enter the URL of the file (Word, PDF, Image, Excel, JSON, or XML): ")
    extracted_text = extract_content(url)
    print("Extracted Content:")
    print(extracted_text)
